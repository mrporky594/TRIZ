import openpyxl

wb = openpyxl.load_workbook(r'C:\TRIZ\TRIZ_Navigator_v5.xlsx', data_only=False)
sheet = wb['TC Matrix & Sample Size']

print(f"Formulas in 'TC Matrix & Sample Size':")
for r in range(1, 15):
    for c in range(1, 26):
        cell = sheet.cell(row=r, column=c)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"Cell {cell.coordinate}: {cell.value}")
        elif cell.value:
            # Also print inputs to understand context
            print(f"Cell {cell.coordinate} (Value): {cell.value}")

print("\nChecking ZScore Lookup for reference values:")
if 'ZScore Lookup' in wb.sheetnames:
    sheet_z = wb['ZScore Lookup']
    for r in range(1, 10):
        row_vals = [sheet_z.cell(row=r, column=c).value for c in range(1, 5)]
        print(f"Row {r}: {row_vals}")
