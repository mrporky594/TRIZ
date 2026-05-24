import openpyxl

wb = openpyxl.load_workbook(r'C:\TRIZ\TRIZ_Navigator_v5.xlsx', data_only=False)
sheets = wb.sheetnames
print(f"Sheets: {sheets}")

for sheet_name in sheets:
    sheet = wb[sheet_name]
    found = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in [1482, 437, 80000000, 345, 6.950000000000001, 6.95]:
                print(f"Sheet '{sheet_name}', Cell {cell.coordinate}: {cell.value}")
                found = True
    if found:
        print("-" * 20)

# Also explicitly check TC Matrix sheet for formulas near the labels
if "TC Matrix" in sheets:
    sheet = wb["TC Matrix"]
    print("\nExploring 'TC Matrix' sheet for formulas:")
    for row in range(1, 25):
        row_values = []
        for col in range(1, 20):
            cell = sheet.cell(row=row, column=col)
            val = cell.value
            if val and isinstance(val, str) and val.startswith('='):
                row_values.append(f"{cell.coordinate}: {val}")
            elif val:
                row_values.append(f"{cell.coordinate}: {val}")
        if row_values:
            print(f"Row {row}: {row_values}")
