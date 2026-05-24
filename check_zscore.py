import openpyxl

wb = openpyxl.load_workbook(r'C:\TRIZ\TRIZ_Navigator_v5.xlsx', data_only=False)
sheet_z = wb['ZScore Lookup']

print("Confidence | Col 2 (Z-score?) | Col 3 (Z-score?)")
for r in range(51, 103):
    conf = sheet_z.cell(row=r, column=1).value
    if conf in [90, 95, 99]:
        v2 = sheet_z.cell(row=r, column=2).value
        v3 = sheet_z.cell(row=r, column=3).value
        print(f"{conf}% | {v2} | {v3}")
