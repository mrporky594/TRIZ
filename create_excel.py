import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_excel():
    try:
        with open('scraped_data_raw.json', 'r') as f:
            data = json.load(f)
    except FileNotFoundError:
        print("Data file not found. Wait for scraper to finish.")
        return

    effects = data.get('effects', {})
    mappings = data.get('mappings', [])

    wb = Workbook()
    
    # --- Sheet 1: Interactive Search ---
    ws_search = wb.active
    ws_search.title = "Search Database"
    headers = ["Mode", "Action / Operation / From", "Object / Parameter / To", "Effect Title", "Description", "Note", "Wikipedia Link"]
    ws_search.append(headers)
    
    # Formatting headers
    for col in range(1, len(headers) + 1):
        cell = ws_search.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for m in mappings:
        mode = m['mode']
        task = m['task_label']
        target = m['target_label']
        for effect_id in m['effect_ids']:
            eff = effects.get(effect_id)
            if eff:
                ws_search.append([
                    mode,
                    task,
                    target,
                    eff['title'],
                    eff['description'],
                    eff['note'],
                    eff['link']
                ])
                row_idx += 1

    # Freeze panes and add filters
    ws_search.freeze_panes = "A2"
    ws_search.auto_filter.ref = f"A1:G{row_idx-1}"

    # Column widths
    widths = [15, 25, 25, 30, 60, 40, 40]
    for i, width in enumerate(widths):
        ws_search.column_dimensions[get_column_letter(i+1)].width = width

    # --- Sheet 2: Master Effects List ---
    ws_effects = wb.create_sheet("All Unique Effects")
    eff_headers = ["ID", "Title", "Description", "Note", "Link"]
    ws_effects.append(eff_headers)
    
    for col in range(1, len(eff_headers) + 1):
        cell = ws_effects.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for eff_id in sorted(effects.keys()):
        eff = effects[eff_id]
        ws_effects.append([eff_id, eff['title'], eff['description'], eff['note'], eff['link']])

    ws_effects.column_dimensions['A'].width = 10
    ws_effects.column_dimensions['B'].width = 30
    ws_effects.column_dimensions['C'].width = 80
    ws_effects.column_dimensions['D'].width = 40
    ws_effects.column_dimensions['E'].width = 40

    # --- Sheet 3: About & Help ---
    ws_about = wb.create_sheet("About & Help")
    ws_about.append(["TRIZ Effects Database - Offline Copy"])
    ws_about.cell(row=1, column=1).font = Font(size=14, bold=True)
    ws_about.append(["This file is a complete offline copy of the TRIZ Effects Database (wbam2244.dns-systems.net)."])
    ws_about.append([])
    ws_about.append(["How to use:"])
    ws_about.append(["1. Go to the 'Search Database' tab."])
    ws_about.append(["2. Use the filter arrows in the column headers to find solutions."])
    ws_about.append(["   - Filter 'Mode' to choose between Function, Parameter, or Transform."])
    ws_about.append(["   - Filter the next two columns to define your query (e.g., 'Move' and 'Liquid')."])
    ws_about.append(["3. The list will instantly update to show all relevant scientific effects."])
    ws_about.append([])
    ws_about.append(["Scraped on: 2026-05-02"])

    wb.save("TRIZ_Effects_Database_Offline.xlsx")
    print("Excel file created: TRIZ_Effects_Database_Offline.xlsx")

if __name__ == "__main__":
    create_excel()
